from __future__ import annotations

import csv
import io
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook

from app.core.exceptions import AppError
from app.modules.analyses.engine.sizing_lab.oracle_engine import (
    STEPS_PER_DAY,
    DayProfile,
    MonthProfile,
)


def _normalize(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _float(value: object) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        number = float(str(value).strip().replace(",", "."))
    except ValueError:
        return None
    return number if np.isfinite(number) else None


def _integer(value: object) -> int | None:
    number = _float(value)
    if number is None or not number.is_integer():
        return None
    return int(number)


def _timestamp(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        pass
    for pattern in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
    ):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def _read_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    return _read_rows_content(path.read_bytes(), path.suffix.lower())


def _read_rows_content(
    content: bytes,
    extension: str,
) -> tuple[list[str], list[dict[str, Any]]]:
    if extension.lower() == ".xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        header = next(iterator, None)
        if header is None:
            workbook.close()
            return [], []
        columns = [str(value or "").strip() for value in header]
        rows = [
            {columns[index]: value for index, value in enumerate(values[: len(columns)])}
            for values in iterator
            if any(value not in (None, "") for value in values)
        ]
        workbook.close()
        return columns, rows

    text: str | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise AppError("Không thể đọc mã hóa CSV.", code="invalid_csv_encoding")
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    columns = [str(item or "").strip() for item in (reader.fieldnames or [])]
    rows = [{str(key or "").strip(): value for key, value in row.items()} for row in reader]
    return columns, rows


def _find_column(columns: list[str], names: set[str]) -> str | None:
    normalized = {_normalize(name) for name in names}
    for column in columns:
        if _normalize(column) in normalized:
            return column
    return None


def _extract_series(dataset: dict[str, Any]) -> tuple[dict[tuple[str, int], float], dict[str, str], bool]:
    source_bytes = dataset.get("source_bytes")
    source_extension = dataset.get("source_extension")
    if isinstance(source_bytes, bytes) and isinstance(source_extension, str):
        columns, rows = _read_rows_content(source_bytes, source_extension)
    else:
        raw_path = dataset.get("source_path")
        if not isinstance(raw_path, str) or not raw_path:
            raise AppError(
                "Dataset không có nguồn dữ liệu để chạy Oracle LP-PF.",
                code="dataset_source_missing",
            )
        path = Path(raw_path)
        if not path.exists():
            raise AppError("File nguồn dataset không tồn tại.", code="dataset_source_missing")
        columns, rows = _read_rows(path)
    timestamp_column = _find_column(
        columns,
        {"timestamp", "time", "datetime", "date_time", "thoi_gian"},
    )
    date_column = _find_column(columns, {"date_iso", "date"})
    day_column = _find_column(columns, {"day_index", "day"})
    step_column = _find_column(
        columns,
        {"step", "time_step", "slot", "interval_index"},
    )
    configured_value = dataset.get("value_column")
    value_column = configured_value if isinstance(configured_value, str) and configured_value in columns else None
    if value_column is None:
        dataset_type = str(dataset.get("dataset_type", ""))
        aliases = (
            {"p_load_kw", "load_kw", "p_load", "load", "demand", "value", "kw", "power"}
            if dataset_type == "load_profile"
            else {"p_pv_kw", "pv_kw", "p_pv", "pv", "pv_power", "solar_kw", "value", "kw", "power"}
        )
        value_column = _find_column(columns, aliases)
    day_type_column = _find_column(columns, {"day_type", "type", "loai_ngay"})
    if value_column is None:
        raise AppError(
            "Không xác định được cột giá trị của dataset để chạy Oracle.",
            code="dataset_value_missing",
            details={"columns": columns},
        )

    values: dict[tuple[str, int], float] = {}
    day_types: dict[str, str] = {}
    has_calendar_dates = False
    if timestamp_column is not None:
        for row in rows:
            ts = _timestamp(row.get(timestamp_column))
            number = _float(row.get(value_column))
            if ts is None or number is None or number < 0:
                continue
            if ts.minute % 15 != 0 or ts.second != 0:
                raise AppError(
                    "Oracle LP-PF yêu cầu dữ liệu đúng bước 15 phút.",
                    code="oracle_interval_required",
                )
            step = ts.hour * 4 + ts.minute // 15
            day_key = ts.date().isoformat()
            values[(day_key, step)] = number
            day_types[day_key] = (
                str(row.get(day_type_column) or "working")
                if day_type_column
                else ("weekend" if ts.weekday() >= 5 else "working")
            )
        has_calendar_dates = True
    elif date_column is not None and step_column is not None:
        for row in rows:
            day_text = str(row.get(date_column) or "").strip()
            step = _integer(row.get(step_column))
            number = _float(row.get(value_column))
            try:
                parsed_date = date.fromisoformat(day_text)
            except ValueError:
                continue
            if step is None or not 0 <= step < STEPS_PER_DAY or number is None or number < 0:
                continue
            day_key = parsed_date.isoformat()
            values[(day_key, step)] = number
            day_types[day_key] = (
                str(row.get(day_type_column) or "working")
                if day_type_column
                else ("weekend" if parsed_date.weekday() >= 5 else "working")
            )
        has_calendar_dates = True
    elif day_column is not None and step_column is not None:
        for row in rows:
            day_index = _integer(row.get(day_column))
            step = _integer(row.get(step_column))
            number = _float(row.get(value_column))
            if day_index is None or day_index < 1 or step is None or not 0 <= step < STEPS_PER_DAY or number is None or number < 0:
                continue
            day_key = str(day_index)
            values[(day_key, step)] = number
            day_types[day_key] = str(row.get(day_type_column) or ("weekend" if (day_index - 1) % 7 >= 5 else "working"))
    else:
        raise AppError(
            "Oracle LP-PF cần timestamp, date_iso + step hoặc day_index + step.",
            code="oracle_time_schema_required",
            details={"columns": columns},
        )
    return values, day_types, has_calendar_dates


def load_profile_months(datasets: list[dict[str, Any]]) -> tuple[list[MonthProfile], dict[str, Any]]:
    load_dataset = next((item for item in datasets if item.get("dataset_type") == "load_profile"), None)
    pv_dataset = next((item for item in datasets if item.get("dataset_type") == "pv_profile"), None)
    if load_dataset is None:
        raise AppError("Cần dataset phụ tải để chạy Sizing Lab.", code="load_dataset_required")

    load_values, load_day_types, dated = _extract_series(load_dataset)
    if not load_values:
        raise AppError("Dataset phụ tải không có dòng hợp lệ.", code="empty_load_profile")
    if pv_dataset is not None:
        pv_values, pv_day_types, pv_dated = _extract_series(pv_dataset)
        if dated != pv_dated:
            raise AppError("Schema thời gian của Load và PV không khớp.", code="dataset_time_mismatch")
    else:
        combined_pv_dataset = {**load_dataset, "dataset_type": "pv_profile", "value_column": None}
        try:
            pv_values, pv_day_types, pv_dated = _extract_series(combined_pv_dataset)
            if dated != pv_dated:
                raise AppError("Schema thời gian của Load và PV không khớp.", code="dataset_time_mismatch")
        except AppError as error:
            if error.code != "dataset_value_missing":
                raise
            pv_values, pv_day_types = {}, {}

    grouped_load: dict[str, dict[int, float]] = defaultdict(dict)
    grouped_pv: dict[str, dict[int, float]] = defaultdict(dict)
    for (day_key, step), value in load_values.items():
        grouped_load[day_key][step] = value
    for (day_key, step), value in pv_values.items():
        grouped_pv[day_key][step] = value

    def sort_key(day_key: str) -> tuple[int, str]:
        return (0, day_key) if dated else (1, f"{int(day_key):09d}")

    days: list[DayProfile] = []
    rejected_days: list[str] = []
    for position, day_key in enumerate(sorted(grouped_load, key=sort_key), start=1):
        load_steps = grouped_load[day_key]
        if len(load_steps) != STEPS_PER_DAY or any(step not in load_steps for step in range(STEPS_PER_DAY)):
            rejected_days.append(day_key)
            continue
        pv_steps = grouped_pv.get(day_key, {})
        pv = np.asarray([pv_steps.get(step, 0.0) for step in range(STEPS_PER_DAY)], dtype=float)
        load = np.asarray([load_steps[step] for step in range(STEPS_PER_DAY)], dtype=float)
        days.append(
            DayProfile(
                load=load,
                pv=pv,
                day_type=load_day_types.get(day_key) or pv_day_types.get(day_key) or "working",
                day_index=position,
                date_iso=day_key if dated else None,
            )
        )
    if not days:
        raise AppError(
            "Không có ngày nào đủ 96 bước 15 phút để chạy Oracle LP-PF.",
            code="oracle_no_complete_days",
        )

    source = str(load_dataset.get("original_name") or load_dataset.get("id") or "uploaded")
    month = MonthProfile(source=source if dated else "block-1", days=days)
    summary = {
        "source": source,
        "n_days": len(days),
        "steps_per_day": STEPS_PER_DAY,
        "interval_minutes": 15,
        "load_peak_kw": round(max(float(day.load.max()) for day in days), 3),
        "load_average_kw": round(float(np.mean(np.concatenate([day.load for day in days]))), 3),
        "pv_peak_kw": round(max(float(day.pv.max()) for day in days), 3),
        "load_energy_kwh": round(sum(float(day.load.sum()) * 0.25 for day in days), 3),
        "pv_energy_kwh": round(sum(float(day.pv.sum()) * 0.25 for day in days), 3),
        "has_calendar_dates": dated,
        "rejected_days": rejected_days,
    }
    return [month], summary
