import asyncio
import csv
import io
import math
import re
import statistics
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from app.core.exceptions import AppError, ConflictError, NotFoundError
from app.dependencies.storage import StorageClient
from app.models.dataset import DatasetDocument
from app.models.file import FileDocument
from app.modules.datasets.enums import DatasetStatus, DatasetType
from app.modules.datasets.repository import DatasetRepository
from app.modules.datasets.schemas import DatasetCreateRequest, DatasetResponse
from app.modules.files.enums import FileStatus
from app.modules.files.repository import FileRepository
from app.modules.projects.repository import ProjectRepository
from app.shared.schemas.pagination import PageMeta, PageResponse

TIMESTAMP_ALIASES = {
    "timestamp",
    "time",
    "datetime",
    "date_time",
    "date",
    "thoi_gian",
    "thoigian",
}
VALUE_ALIASES = {
    "value",
    "kw",
    "kwh",
    "power",
    "energy",
    "load",
    "demand",
    "cong_suat",
    "congsuat",
    "dien_nang",
    "diennang",
    "pv",
    "pv_power",
}
LOAD_VALUE_ALIASES = VALUE_ALIASES | {
    "p_load_kw",
    "load_kw",
    "p_load",
    "active_load_kw",
}
PV_VALUE_ALIASES = VALUE_ALIASES | {
    "p_pv_kw",
    "pv_kw",
    "p_pv",
    "solar_kw",
    "solar_power_kw",
}
DAY_INDEX_ALIASES = {"day_index", "day", "ngay_index", "ngay"}
DATE_INDEX_ALIASES = {"date_iso", "date"}
STEP_ALIASES = {"step", "time_step", "slot", "interval_index", "buoc"}
EMS_STEPS_PER_DAY = 96
EMS_INTERVAL_MINUTES = 15


class DatasetService:
    def __init__(
        self,
        dataset_repository: DatasetRepository,
        file_repository: FileRepository,
        project_repository: ProjectRepository,
        storage_client: StorageClient,
    ) -> None:
        self.dataset_repository = dataset_repository
        self.file_repository = file_repository
        self.project_repository = project_repository
        self.storage_client = storage_client

    async def create_dataset(
        self,
        payload: DatasetCreateRequest,
        user_id: str,
    ) -> DatasetResponse:
        project = await self.project_repository.get_by_id_for_user(payload.project_id, user_id)
        if project is None:
            raise NotFoundError("Project not found.")
        file_document = await self.file_repository.get_by_id_for_user(payload.file_id, user_id)
        if file_document is None or file_document.project_id != payload.project_id:
            raise NotFoundError("Project file not found.")
        if file_document.kind.value != payload.dataset_type.value:
            raise AppError(
                "File kind does not match the requested dataset type.",
                code="dataset_type_mismatch",
            )
        existing = await self.dataset_repository.get_by_file_for_user(payload.file_id, user_id)
        if existing is not None:
            if existing.project_id == payload.project_id and existing.dataset_type == payload.dataset_type:
                return self._to_response(existing)
            raise ConflictError("This file is already linked to another dataset.")

        exists = await asyncio.to_thread(
            self.storage_client.exists,
            file_document.storage_path,
        )
        if not exists:
            raise NotFoundError("Stored file is missing.")

        parsed = await asyncio.to_thread(
            _parse_stored_source,
            self.storage_client,
            file_document,
            payload.dataset_type,
        )
        dataset = DatasetDocument(
            user_id=user_id,
            project_id=payload.project_id,
            file_id=payload.file_id,
            dataset_type=payload.dataset_type,
            status=parsed["status"],
            row_count=parsed["row_count"],
            valid_row_count=parsed["valid_row_count"],
            interval_minutes=parsed["interval_minutes"],
            columns=parsed["columns"],
            timestamp_column=parsed["timestamp_column"],
            value_column=parsed["value_column"],
            start_at=parsed["start_at"],
            end_at=parsed["end_at"],
            quality_summary=parsed["quality_summary"],
            preview=parsed["preview"],
        )
        created = await self.dataset_repository.create_dataset(dataset)
        if created.id is None:
            raise AppError("Dataset was created without an identifier.", code="dataset_id_missing")
        updated_project = await self.project_repository.add_dataset_id_for_user(
            payload.project_id,
            user_id,
            created.id,
        )
        if updated_project is None:
            await self.dataset_repository.delete_by_id_for_user(created.id, user_id)
            raise NotFoundError("Project not found while linking dataset.")
        await self.file_repository.update_by_id_for_user(
            payload.file_id,
            user_id,
            {
                "status": (
                    FileStatus.INVALID
                    if created.status == DatasetStatus.INVALID
                    else FileStatus.VALIDATED
                ),
                "metadata": {
                    "dataset_id": created.id,
                    "dataset_type": created.dataset_type,
                    "row_count": created.row_count,
                    "valid_row_count": created.valid_row_count,
                    "interval_minutes": created.interval_minutes,
                },
            },
        )
        return self._to_response(created)

    async def list_datasets(
        self,
        user_id: str,
        *,
        page: int,
        page_size: int,
        skip: int,
        project_id: str | None = None,
    ) -> PageResponse[DatasetResponse]:
        total = await self.dataset_repository.count_by_user(user_id, project_id)
        datasets = await self.dataset_repository.list_by_user(
            user_id,
            skip=skip,
            limit=page_size,
            project_id=project_id,
        )
        return PageResponse[DatasetResponse](
            items=[self._to_response(item) for item in datasets],
            meta=PageMeta(page=page, page_size=page_size, total=total),
        )

    async def get_dataset(self, dataset_id: str, user_id: str) -> DatasetResponse:
        dataset = await self.dataset_repository.get_by_id_for_user(dataset_id, user_id)
        if dataset is None:
            raise NotFoundError("Dataset not found.")
        return self._to_response(dataset)

    async def delete_dataset(self, dataset_id: str, user_id: str) -> None:
        dataset = await self.dataset_repository.delete_by_id_for_user(dataset_id, user_id)
        if dataset is None:
            raise NotFoundError("Dataset not found.")
        await self.project_repository.remove_dataset_id_for_user(
            dataset.project_id,
            user_id,
            dataset_id,
        )
        await self.file_repository.update_by_id_for_user(
            dataset.file_id,
            user_id,
            {"status": FileStatus.UPLOADED, "metadata": {}},
        )

    @staticmethod
    def _to_response(dataset: DatasetDocument) -> DatasetResponse:
        return DatasetResponse.model_validate(dataset)


def _parse_stored_source(
    storage_client: StorageClient,
    file_document: FileDocument,
    dataset_type: DatasetType,
) -> dict[str, Any]:
    with storage_client.materialize(file_document.storage_path) as path:
        return _parse_source(path, file_document, dataset_type)


def _parse_source(
    path: Path,
    file_document: FileDocument,
    dataset_type: DatasetType,
) -> dict[str, Any]:
    extension = file_document.extension.lower()
    if extension == "csv":
        columns, rows = _read_csv(path)
    elif extension == "xlsx":
        columns, rows = _read_xlsx(path)
    else:
        raise AppError("Unsupported dataset file type.", code="unsupported_dataset_type")
    return _analyze_rows(columns, rows, dataset_type=dataset_type)


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    content = path.read_bytes()
    text: str | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise AppError("Could not decode CSV file.", code="invalid_csv_encoding")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    columns = [str(item or "").strip() for item in (reader.fieldnames or [])]
    rows = [{str(key or "").strip(): value for key, value in row.items()} for row in reader]
    return columns, rows


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    header_values = next(iterator, None)
    if header_values is None:
        workbook.close()
        return [], []
    columns = [str(value).strip() if value is not None else "" for value in header_values]
    rows: list[dict[str, Any]] = []
    for values in iterator:
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({columns[index]: value for index, value in enumerate(values[: len(columns)])})
    workbook.close()
    return columns, rows


def _analyze_rows(
    columns: list[str],
    rows: list[dict[str, Any]],
    *,
    dataset_type: DatasetType | None = None,
) -> dict[str, Any]:
    if len(columns) < 2:
        raise AppError("Dataset requires at least two columns.", code="invalid_dataset_columns")

    timestamp_column = _find_column(columns, TIMESTAMP_ALIASES)
    date_index_column = _find_column(columns, DATE_INDEX_ALIASES)
    day_index_column = _find_column(columns, DAY_INDEX_ALIASES)
    step_column = _find_column(columns, STEP_ALIASES)
    if (
        timestamp_column is not None
        and date_index_column == timestamp_column
        and step_column is not None
    ):
        timestamp_column = None
    calendar_indexed_schema = (
        timestamp_column is None
        and date_index_column is not None
        and step_column is not None
    )
    ordinal_indexed_schema = (
        timestamp_column is None
        and day_index_column is not None
        and step_column is not None
    )
    indexed_schema = calendar_indexed_schema or ordinal_indexed_schema

    value_aliases = VALUE_ALIASES
    if dataset_type == DatasetType.LOAD_PROFILE:
        value_aliases = LOAD_VALUE_ALIASES
    elif dataset_type == DatasetType.PV_PROFILE:
        value_aliases = PV_VALUE_ALIASES
    value_column = _find_column(
        columns,
        value_aliases,
        excluded={
            timestamp_column,
            date_index_column,
            day_index_column,
            step_column,
        },
    )
    if (timestamp_column is None and not indexed_schema) or value_column is None:
        raise AppError(
            "Could not identify a supported time schema and power/energy column.",
            code="dataset_columns_not_found",
            details={
                "columns": columns,
                "supported_time_schemas": [
                    "timestamp",
                    "date_iso + step (96 steps/day, 15 minutes)",
                    "day_index + step (96 steps/day, 15 minutes)",
                ],
            },
        )

    parsed_rows: list[tuple[datetime, float]] = []
    invalid_rows = 0
    negative_values = 0
    indexed_steps: dict[str | int, set[int]] = {}
    preview: list[dict[str, Any]] = []
    for row in rows:
        if calendar_indexed_schema:
            date_value = _parse_date(row.get(date_index_column))
            step = _parse_integer(row.get(step_column))
            timestamp = _parse_date_indexed_timestamp(date_value, step)
            if date_value is not None and step is not None:
                indexed_steps.setdefault(date_value.isoformat(), set()).add(step)
        elif ordinal_indexed_schema:
            day_index = _parse_integer(row.get(day_index_column))
            step = _parse_integer(row.get(step_column))
            timestamp = _parse_indexed_timestamp(day_index, step)
            if day_index is not None and step is not None:
                indexed_steps.setdefault(day_index, set()).add(step)
        else:
            timestamp = _parse_timestamp(row.get(timestamp_column))
        value = _parse_number(row.get(value_column))
        if timestamp is None or value is None:
            invalid_rows += 1
            continue
        if value < 0:
            negative_values += 1
        parsed_rows.append((timestamp, value))
        if len(preview) < 10:
            preview.append(
                {
                    "timestamp": timestamp.isoformat(),
                    "value": round(value, 6),
                }
            )

    parsed_rows.sort(key=lambda item: item[0])
    timestamp_counts: dict[datetime, int] = {}
    for timestamp, _ in parsed_rows:
        timestamp_counts[timestamp] = timestamp_counts.get(timestamp, 0) + 1
    duplicate_timestamps = sum(count - 1 for count in timestamp_counts.values() if count > 1)
    unique_timestamps = sorted(timestamp_counts)
    intervals = [
        (right - left).total_seconds() / 60
        for left, right in zip(unique_timestamps, unique_timestamps[1:])
        if right > left
    ]
    interval_minutes = (
        float(EMS_INTERVAL_MINUTES)
        if indexed_schema
        else round(statistics.median(intervals), 3) if intervals else None
    )
    irregular_intervals = 0
    if interval_minutes and intervals:
        tolerance = max(0.01, interval_minutes * 0.05)
        irregular_intervals = sum(
            1 for interval in intervals if abs(interval - interval_minutes) > tolerance
        )
    incomplete_days = (
        sum(1 for steps in indexed_steps.values() if len(steps) != EMS_STEPS_PER_DAY)
        if indexed_schema
        else 0
    )

    warnings: list[str] = []
    if invalid_rows:
        warnings.append(f"{invalid_rows} rows could not be parsed.")
    if duplicate_timestamps:
        warnings.append(f"{duplicate_timestamps} duplicate timestamps detected.")
    if irregular_intervals:
        warnings.append(f"{irregular_intervals} irregular intervals detected.")
    if incomplete_days:
        warnings.append(
            f"{incomplete_days} indexed days do not contain exactly {EMS_STEPS_PER_DAY} steps."
        )
    if negative_values:
        warnings.append(f"{negative_values} negative values detected.")

    valid_row_count = len(parsed_rows)
    parsed_values = [value for _, value in parsed_rows]
    observed_hours = (
        valid_row_count * interval_minutes / 60
        if interval_minutes is not None
        else None
    )
    status = (
        DatasetStatus.INVALID
        if valid_row_count == 0
        else DatasetStatus.WARNING
        if warnings
        else DatasetStatus.READY
    )
    resolved_timestamp_column = timestamp_column
    if calendar_indexed_schema:
        resolved_timestamp_column = f"{date_index_column}+{step_column}"
    elif ordinal_indexed_schema:
        resolved_timestamp_column = f"{day_index_column}+{step_column}"
    return {
        "status": status,
        "row_count": len(rows),
        "valid_row_count": valid_row_count,
        "interval_minutes": interval_minutes,
        "columns": columns,
        "timestamp_column": resolved_timestamp_column,
        "value_column": value_column,
        "start_at": parsed_rows[0][0] if parsed_rows else None,
        "end_at": parsed_rows[-1][0] if parsed_rows else None,
        "preview": preview,
        "quality_summary": {
            "source_schema": (
                "date_iso_step"
                if calendar_indexed_schema
                else "day_index_step"
                if ordinal_indexed_schema
                else "timestamp"
            ),
            "date_index_column": date_index_column if calendar_indexed_schema else None,
            "day_index_column": day_index_column if ordinal_indexed_schema else None,
            "step_column": step_column if indexed_schema else None,
            "indexed_days": len(indexed_steps) if indexed_schema else None,
            "incomplete_days": incomplete_days,
            "invalid_rows": invalid_rows,
            "duplicate_timestamps": duplicate_timestamps,
            "irregular_intervals": irregular_intervals,
            "negative_values": negative_values,
            "value_min": min(parsed_values) if parsed_values else None,
            "value_max": max(parsed_values) if parsed_values else None,
            "value_mean": statistics.fmean(parsed_values) if parsed_values else None,
            "value_sum": sum(parsed_values) if parsed_values else None,
            "observed_hours": observed_hours,
            "warnings": warnings,
        },
    }


def _find_column(
    columns: Iterable[str],
    aliases: set[str],
    *,
    excluded: set[str | None] | None = None,
) -> str | None:
    excluded = excluded or set()
    normalized_aliases = {_normalize_name(alias) for alias in aliases}
    for column in columns:
        if column in excluded:
            continue
        if _normalize_name(column) in normalized_aliases:
            return column
    for column in columns:
        if column in excluded:
            continue
        normalized = _normalize_name(column)
        if any(alias in normalized for alias in normalized_aliases):
            return column
    return None


def _normalize_name(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value.strip().lower())
    without_accents = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
    return re.sub(r"[^a-z0-9]+", "_", without_accents).strip("_")


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _parse_integer(value: Any) -> int | None:
    number = _parse_number(value)
    if number is None or not number.is_integer():
        return None
    return int(number)


def _parse_date_indexed_timestamp(
    date_value: date | None,
    step: int | None,
) -> datetime | None:
    if date_value is None or step is None or step < 0 or step >= EMS_STEPS_PER_DAY:
        return None
    return datetime.combine(date_value, datetime.min.time()) + timedelta(
        minutes=step * EMS_INTERVAL_MINUTES,
    )


def _parse_indexed_timestamp(day_index: int | None, step: int | None) -> datetime | None:
    if day_index is None or step is None:
        return None
    if day_index < 1 or step < 0 or step >= EMS_STEPS_PER_DAY:
        return None
    return datetime(2000, 1, 1) + timedelta(
        days=day_index - 1,
        minutes=step * EMS_INTERVAL_MINUTES,
    )


def _parse_timestamp(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        pass
    for format_string in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
    ):
        try:
            return datetime.strptime(text, format_string)
        except ValueError:
            continue
    return None


def _parse_number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        text = "".join(parts) if len(parts[-1]) == 3 else text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None
